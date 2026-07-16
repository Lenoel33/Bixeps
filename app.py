from __future__ import annotations

import io
import re
from pathlib import Path
import gc
import html

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from analyser import analyse_workbook, build_pdf_report, find_header_row

st.set_page_config(
    page_title="BIXEPS Assessment Analyser",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
.block-container {padding-top: 1.6rem; padding-bottom: 3rem;}
[data-testid="stMetric"] {background: #ffffff; border: 1px solid #e6e9ef; padding: 14px; border-radius: 12px;}
[data-testid="stMetricLabel"] {font-weight: 650; min-height: 3.1rem; align-items: flex-start;}
[data-testid="stMetricLabel"] > div {white-space: normal !important; overflow: visible !important; text-overflow: clip !important; line-height: 1.25 !important;}
[data-testid="stMetricLabel"] p {white-space: normal !important; overflow: visible !important; text-overflow: clip !important; line-height: 1.25 !important;}
[data-testid="stMetricValue"] {line-height: 1.15;}
[data-testid="stMetricDelta"] {white-space: normal !important; overflow: visible !important; text-overflow: clip !important;}
.small-note {color:#5f6b7a; font-size:0.92rem;}
.section-title {font-size:1.25rem; font-weight:700; margin-top:0.4rem;}
.wrapped-table-container {
    width: 100%;
    max-height: 600px;
    overflow: auto;
    border: 1px solid #e3e7ed;
    border-radius: 10px;
    background: white;
}
.wrapped-table {
    border-collapse: separate;
    border-spacing: 0;
    width: max-content;
    min-width: 100%;
    font-size: 0.93rem;
}
.wrapped-table th {
    position: sticky;
    top: 0;
    z-index: 2;
    background: #f4f6f8;
    color: #2f3542;
    font-weight: 700;
    text-align: left;
    border-bottom: 1px solid #dfe3e8;
    padding: 10px 12px;
    white-space: normal;
}
.wrapped-table td {
    vertical-align: top;
    padding: 10px 12px;
    border-bottom: 1px solid #edf0f3;
    white-space: normal !important;
    overflow-wrap: anywhere;
    word-break: normal;
    line-height: 1.45;
}
.wrapped-table tr:nth-child(even) td {background: #fafbfc;}
.wrapped-table tr:hover td {background: #f5f9ff;}
.wrapped-table .short-col {min-width: 105px; max-width: 150px;}
.wrapped-table .medium-col {min-width: 170px; max-width: 260px;}
.wrapped-table .long-col {min-width: 360px; max-width: 620px;}
</style>
""",
    unsafe_allow_html=True,
)

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

def clear_uploaded_data() -> None:
    """Remove uploaded data from the active Streamlit session and rerun cleanly."""
    st.session_state.uploader_key += 1
    gc.collect()


def render_wrapped_table(
    dataframe: pd.DataFrame,
    *,
    height: int = 560,
    percentage_columns: tuple[str, ...] = (),
) -> None:
    """Render a scrollable HTML table with fully wrapped text.

    Streamlit's standard dataframe grid truncates long paragraph fields. This
    renderer escapes workbook content, wraps every cell, and gives narrative
    columns additional width so the complete reasons remain readable.
    """
    table_df = dataframe.copy()
    for column in percentage_columns:
        if column in table_df.columns:
            table_df[column] = table_df[column].apply(
                lambda value: "" if pd.isna(value) else f"{float(value):.0%}"
            )

    long_keywords = (
        "reason", "evidence", "remark", "recommendation",
        "explanation", "follow-up",
    )
    medium_keywords = (
        "transition", "domain", "outcome", "status", "risk",
    )

    headers = []
    for column in table_df.columns:
        lowered = str(column).casefold()
        if any(keyword in lowered for keyword in long_keywords):
            css_class = "long-col"
        elif any(keyword in lowered for keyword in medium_keywords):
            css_class = "medium-col"
        else:
            css_class = "short-col"
        headers.append(f'<th class="{css_class}">{html.escape(str(column))}</th>')

    rows = []
    for _, row in table_df.iterrows():
        cells = []
        for column, value in row.items():
            lowered = str(column).casefold()
            if any(keyword in lowered for keyword in long_keywords):
                css_class = "long-col"
            elif any(keyword in lowered for keyword in medium_keywords):
                css_class = "medium-col"
            else:
                css_class = "short-col"
            if pd.isna(value):
                display_value = ""
            else:
                display_value = html.escape(str(value)).replace("\n", "<br>")
            cells.append(f'<td class="{css_class}">{display_value}</td>')
        rows.append("<tr>" + "".join(cells) + "</tr>")

    table_html = (
        f'<div class="wrapped-table-container" style="max-height:{height}px">'
        '<table class="wrapped-table"><thead><tr>'
        + "".join(headers)
        + "</tr></thead><tbody>"
        + "".join(rows)
        + "</tbody></table></div>"
    )
    st.markdown(table_html, unsafe_allow_html=True)

st.title("📊 BIXEPS Assessment Analyser")
st.caption(
    "Upload a completed BIXEPS Excel workbook to automatically identify improvements, stable outcomes, declines and seniors requiring follow-up."
)

with st.sidebar:
    st.header("Analysis settings")
    st.write("The analyser automatically detects the assessment sheet and keeps every original workbook sheet unchanged.")
    show_incomplete = st.checkbox("Include incomplete assessments in tables", value=True)
    follow_up_mode = st.selectbox(
        "Follow-up priority",
        ["Standard", "Sensitive"],
        help="Sensitive mode also flags any senior with worsening SPPB, mobility or pain even when the overall result is maintained.",
    )
    st.divider()
    st.markdown("**Privacy controls**")
    st.caption("Files are processed in memory only. The app does not save uploads, generated reports, or senior records to disk or a database.")
    if st.button("Clear uploaded data", use_container_width=True):
        clear_uploaded_data()
        st.rerun()
    st.divider()
    st.markdown("**Result meanings**")
    st.markdown("🟢 Improved  \n🔵 Maintained  \n⚪ No Change  \n🔴 Declined  \n🟡 Insufficient Data")

uploaded = st.file_uploader("Upload BIXEPS Excel file", type=["xlsx"], help="Use the original completed assessment workbook.", key=f"bixeps_upload_{st.session_state.uploader_key}")

if not uploaded:
    st.info("Upload an Excel workbook to begin. Uploaded data is processed in memory and is not saved by this app.")
    with st.expander("What the analyser produces", expanded=True):
        st.markdown(
            """
- A management dashboard with outcome counts and percentages
- Average SPPB change, pain reduction and health-score change
- Frailty-status improvement statistics
- Seniors requiring follow-up and the reason for each flag
- Top improving seniors
- A downloadable analysed Excel workbook and PDF management report
            """
        )
    st.stop()

raw = uploaded.getvalue()
try:
    preview_wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    candidate_sheets: list[str] = []
    for name in preview_wb.sheetnames:
        try:
            find_header_row(preview_wb[name])
            candidate_sheets.append(name)
        except ValueError:
            continue

    if not candidate_sheets:
        st.error("No sheet containing the required BIXEPS headings was found.")
        st.stop()

    selected_sheet = candidate_sheets[0]
    if len(candidate_sheets) > 1:
        selected_sheet = st.selectbox("Select assessment data sheet", candidate_sheets)

    with st.spinner("Analysing assessment results..."):
        df, result_bytes, used_sheet = analyse_workbook(raw, selected_sheet)
        pdf_bytes = build_pdf_report(df, Path(uploaded.name).stem)

    st.success(f"Successfully analysed {len(df)} seniors from '{used_sheet}'.")
    st.caption("Privacy: the workbook and generated reports exist only in this active session. Use ‘Clear uploaded data’ when finished, then close the browser tab.")

    complete_df = df[df["Overall Outcome"] != "Insufficient Data"].copy()
    counts = df["Overall Outcome"].value_counts()

    st.markdown('<div class="section-title">Assessment status</div>', unsafe_allow_html=True)
    k1, k2, k3 = st.columns(3)
    k1.metric("Seniors analysed", len(df))
    k2.metric("Improved", int(counts.get("Improved", 0)))
    k3.metric("Maintained", int(counts.get("Maintained", 0)))
    k4, k5, k6 = st.columns(3)
    k4.metric("No Change", int(counts.get("No Change", 0)))
    k5.metric("Declined", int(counts.get("Declined", 0)))
    k6.metric("Incomplete assessments", int(counts.get("Insufficient Data", 0)))

    tabs = st.tabs(["Overview", "Frailty Impact", "Remarks Review", "Follow-up", "Top Improvers", "All Seniors", "Downloads"])

    with tabs[0]:
        st.markdown('<div class="section-title">Programme outcome overview</div>', unsafe_allow_html=True)
        left, right = st.columns([1, 1])
        outcome_order = ["Improved", "Maintained", "No Change", "Declined", "Insufficient Data"]
        outcome_chart = pd.DataFrame({"Outcome": outcome_order, "Seniors": [int(counts.get(x, 0)) for x in outcome_order]})
        with left:
            st.bar_chart(outcome_chart.set_index("Outcome"), y="Seniors", use_container_width=True)
        with right:
            denominator = max(len(complete_df), 1)
            summary = pd.DataFrame(
                {
                    "Indicator": ["Improved", "Maintained", "No Change", "Declined"],
                    "Percentage": [
                        counts.get("Improved", 0) / denominator,
                        counts.get("Maintained", 0) / denominator,
                        counts.get("No Change", 0) / denominator,
                        counts.get("Declined", 0) / denominator,
                    ],
                }
            )
            st.dataframe(
                summary.style.format({"Percentage": "{:.1%}"}),
                use_container_width=True,
                hide_index=True,
            )
            st.caption("Percentages exclude seniors with insufficient post-assessment data.")

        st.markdown('<div class="section-title">Programme impact indicators</div>', unsafe_allow_html=True)
        m1, m2, m3 = st.columns(3)
        m1.metric("Average SPPB score change", f"{complete_df['SPPB Change'].mean():+.2f}" if not complete_df.empty else "N/A")
        m2.metric(
            "Average overall pain score change",
            f"{complete_df['Overall Pain Change'].mean():+.2f}" if not complete_df.empty else "N/A",
            help="A negative value means that average pain reduced.",
        )
        m3.metric(
            "Average overall health score change",
            f"{complete_df['Health Score Change'].mean():+.2f}" if not complete_df.empty else "N/A",
        )
        frailty_available = df["Frailty Outcome"].isin(["Improved", "No Change", "Declined"])
        frailty_improved = int((df.loc[frailty_available, "Frailty Outcome"] == "Improved").sum())
        frailty_total = int(frailty_available.sum())
        recovered_count = int(df["Recovered from Frailty"].sum())
        verified_recovered = int(((df["Recovered from Frailty"]) & df["SPPB Change"].notna()).sum())
        pre_frail_complete = df[
            df["Frailty Status"].astype(str).str.casefold().str.contains("frail", na=False)
            & df["Frailty Status 2"].notna()
        ]
        recovery_rate = verified_recovered / len(pre_frail_complete) if len(pre_frail_complete) else 0

        m4, m5 = st.columns(2)
        m4.metric(
            "Seniors with improved frailty status",
            f"{frailty_improved} of {frailty_total}" if frailty_total else "N/A",
        )
        m5.metric(
            "Verified transitions from frail to normal",
            verified_recovered,
            delta=f"{recovery_rate:.1%} of seniors who were frail before BIXEPS",
        )

        domain_cols = ["Physical Function Domain", "Pain Domain", "Daily Living Domain", "Wellbeing Domain"]
        domain_rows = []
        for col in domain_cols:
            eligible = df[df[col].isin(["Improved", "Maintained", "No Change", "Declined"])]
            denominator = len(eligible)
            improved_n = int((eligible[col] == "Improved").sum())
            stable_n = int(eligible[col].isin(["Maintained", "No Change"]).sum())
            declined_n = int((eligible[col] == "Declined").sum())
            domain_rows.append(
                {
                    "Domain": col.replace(" Domain", ""),
                    "Improved %": improved_n / denominator if denominator else 0,
                    "Maintained / No Change %": stable_n / denominator if denominator else 0,
                    "Declined %": declined_n / denominator if denominator else 0,
                    "Complete assessments (n)": denominator,
                }
            )
        st.markdown('<div class="section-title">Results by assessment domain</div>', unsafe_allow_html=True)
        domain_table = pd.DataFrame(domain_rows)
        st.dataframe(
            domain_table.style.format({
                "Improved %": "{:.1%}",
                "Maintained / No Change %": "{:.1%}",
                "Declined %": "{:.1%}",
            }),
            use_container_width=True,
            hide_index=True,
        )
        st.caption("Each percentage uses only seniors with enough comparable data for that domain. The denominator is shown in the final column.")

    with tabs[1]:
        st.markdown('<div class="section-title">Frailty improvement after BIXEPS</div>', unsafe_allow_html=True)
        st.info("This section shows improvement observed after participation. It supports programme effectiveness, but does not by itself prove that BIXEPS caused every change because there is no comparison group.")
        recovered = df[df["Recovered from Frailty"]].copy()
        pre_frail = df[df["Frailty Status"].astype(str).str.casefold().str.contains("frail", na=False)]
        pre_frail_complete = pre_frail[pre_frail["Frailty Status 2"].notna()]
        f1, f2, f3 = st.columns(3)
        f1.metric("Pre-assessment frail", len(pre_frail_complete))
        verified = recovered[recovered["SPPB Change"].notna()]
        f2.metric("Verified improvement to normal", len(verified))
        f3.metric("Verified frail-to-normal rate", f"{len(verified) / len(pre_frail_complete):.1%}" if len(pre_frail_complete) else "N/A")
        if recovered.empty:
            st.warning("No frail-to-normal transitions were detected in this file.")
        else:
            st.success(f"{len(verified)} verified and {len(recovered) - len(verified)} provisional frail-to-normal transition(s) were detected.")
            st.caption("Verified cases have complete pre/post SPPB scores. Provisional cases are retained but clearly flagged because their SPPB evidence is incomplete.")
            frailty_cols = [
                "Name", "Gender", "Age (in 2026)", "Frailty Transition", "Frailty Evidence", "SPPB Pre", "SPPB Post",
                "SPPB Change", "Overall Outcome", "Analysis Reason",
            ]
            render_wrapped_table(recovered[frailty_cols], height=440)
            st.download_button(
                "Download frailty improvement list (CSV)",
                recovered[frailty_cols].to_csv(index=False).encode("utf-8-sig"),
                file_name="BIXEPS_Frail_to_Normal.csv",
                mime="text/csv",
            )

    with tabs[2]:
        st.markdown('<div class="section-title">Context review of written remarks</div>', unsafe_allow_html=True)
        st.caption("The system uses rule-based contextual checks. It distinguishes negated entries such as ‘No numbness’ from persistent or frequent symptoms such as ‘still numb’ or ‘2× a week’. It flags cases for staff review and does not make a medical diagnosis.")
        text_counts = df["Text Risk Level"].value_counts()
        r1, r2, r3 = st.columns(3)
        r1.metric("Priority review", int(text_counts.get("Escalate", 0)))
        r2.metric("Monitor", int(text_counts.get("Monitor", 0)))
        r3.metric("No concern detected", int(text_counts.get("No concern detected", 0)))
        text_review = df[df["Text Risk Level"].isin(["Escalate", "Monitor"])][[
            "Name", "Text Risk Level", "Text Risk Reason", "Text Evidence", "Overall Outcome", "Frailty Transition"
        ]]
        if text_review.empty:
            st.success("No written remarks were flagged for review.")
        else:
            render_wrapped_table(text_review, height=480)
            st.download_button(
                "Download remarks review list (CSV)",
                text_review.to_csv(index=False).encode("utf-8-sig"),
                file_name="BIXEPS_Remarks_Review.csv",
                mime="text/csv",
            )

    with tabs[3]:
        st.markdown('<div class="section-title">Seniors requiring review</div>', unsafe_allow_html=True)
        if follow_up_mode == "Sensitive":
            follow_mask = (df["Overall Outcome"] == "Declined") | df["Critical Decline"] | df["Text Risk Level"].isin(["Escalate", "Monitor"])
        else:
            follow_mask = (df["Overall Outcome"] == "Declined") | (df["Text Risk Level"] == "Escalate")
        follow_df = df[follow_mask].copy()
        st.metric("Seniors flagged", len(follow_df))
        if follow_df.empty:
            st.success("No seniors meet the selected follow-up criteria.")
        else:
            cols = [
                "Name", "Gender", "Age (in 2026)", "Overall Outcome", "SPPB Outcome",
                "Overall Pain Outcome", "Mobility Outcome", "Frailty Outcome", "Frailty Transition",
                "Text Risk Level", "Text Risk Reason", "Text Evidence", "Follow-up Reason", "Analysis Reason",
            ]
            render_wrapped_table(follow_df[cols], height=560)
            st.download_button(
                "Download follow-up list (CSV)",
                follow_df[cols].to_csv(index=False).encode("utf-8-sig"),
                file_name="BIXEPS_Follow_Up_List.csv",
                mime="text/csv",
            )

    with tabs[4]:
        st.markdown('<div class="section-title">Top improving seniors</div>', unsafe_allow_html=True)
        ranked = df[df["Overall Outcome"] != "Insufficient Data"].sort_values(
            ["Improvement Score", "Improved Areas", "SPPB Change"], ascending=[False, False, False]
        ).head(10)
        top_cols = [
            "Name", "Overall Outcome", "Improvement Score", "Improved Areas", "Declined Areas",
            "SPPB Change", "Overall Pain Change", "Health Score Change", "Analysis Reason",
        ]
        render_wrapped_table(ranked[top_cols], height=520)

    with tabs[5]:
        st.markdown('<div class="section-title">Complete senior-level analysis with reasons</div>', unsafe_allow_html=True)
        view_df = df if show_incomplete else df[df["Overall Outcome"] != "Insufficient Data"]
        outcome_filter = st.multiselect(
            "Filter by overall outcome",
            ["Improved", "Maintained", "No Change", "Declined", "Insufficient Data"],
            default=["Improved", "Maintained", "No Change", "Declined", "Insufficient Data"] if show_incomplete else ["Improved", "Maintained", "No Change", "Declined"],
        )
        view_df = view_df[view_df["Overall Outcome"].isin(outcome_filter)]
        display_columns = [
            "Name", "Gender", "Age (in 2026)", "Overall Outcome", "Frailty Transition",
            "SPPB Outcome", "Physical Function Domain", "Pain Domain", "Daily Living Domain", "Wellbeing Domain",
            "Improved Areas", "Declined Areas", "Data Completeness", "Text Risk Level", "Text Risk Reason", "Text Evidence",
            "Analysis Reason", "Physical Function Reason", "Pain Reason", "Daily Living Reason", "Wellbeing Reason",
        ]
        render_wrapped_table(
            view_df[display_columns],
            height=620,
            percentage_columns=("Data Completeness",),
        )
        st.caption("All narrative fields wrap automatically. Scroll horizontally for additional columns and vertically for more seniors.")

    with tabs[6]:
        st.markdown('<div class="section-title">Download reports</div>', unsafe_allow_html=True)
        output_name = re.sub(r"\.xlsx$", "", uploaded.name, flags=re.I) + "_Analysed.xlsx"
        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "⬇ Download analysed Excel workbook",
                data=result_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
            st.caption("Contains the original sheets plus Analysis Dashboard, Senior Analysis, Frailty Improvements, Follow-up List and Top Improvers.")
        with d2:
            st.download_button(
                "⬇ Download management PDF report",
                data=pdf_bytes,
                file_name=re.sub(r"\.xlsx$", "", uploaded.name, flags=re.I) + "_Management_Report.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
            st.caption("A concise report of programme outcomes, key changes and seniors requiring review.")

except Exception as exc:
    st.error(f"The workbook could not be analysed: {exc}")
    st.exception(exc)
