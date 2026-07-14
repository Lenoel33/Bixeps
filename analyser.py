from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Any, Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


@dataclass(frozen=True)
class Metric:
    label: str
    pre_header: str
    post_header: str
    direction: str
    weight: float = 1.0
    domain: str = "Other"


METRICS = [
    Metric("SPPB", "SPPB Overall Score", "SPPB Overall Score 2", "higher", 2.0, "Physical Function"),
    Metric("Overall Pain", "Pre Bixeps Overall VAS", "Post BIXEPS Overall VAS", "lower", 1.5, "Pain"),
    Metric("Lower-Limb Pain", "Pre Bixeps LL VAS", "Post BIXEPS LL VAS", "lower", 1.5, "Pain"),
    Metric("Other Pain", "Pre Bixeps Other VAS", "Post BIXEPS Other VAS", "lower", 1.0, "Pain"),
    Metric("Mobility", "Pre-BIXEPS", "Post-BIXEPS", "lower", 1.5, "Daily Living"),
    Metric("Self-Care", "Pre-BIXEPS 2", "Post-BIXEPS 2", "lower", 1.0, "Daily Living"),
    Metric("Usual Activities", "Pre-BIXEPS 3", "Post-BIXEPS 3", "lower", 1.0, "Daily Living"),
    Metric("Pain/Discomfort", "Pre-BIXEPS 4", "Post-BIXEPS 4", "lower", 1.0, "Wellbeing"),
    Metric("Anxiety/Depression", "Pre-BIXEPS 5", "Post-BIXEPS 5", "lower", 1.0, "Wellbeing"),
    Metric("Health Questionnaire", "Pre-BIXEPS 6", "Post-BIXEPS 6", "lower", 1.5, "Wellbeing"),
    Metric("Numbness in Feet", "Pre-BIXEPS 7", "Post-BIXEPS 7", "symptom", 1.0, "Wellbeing"),
    Metric("Night Leg Cramps", "Pre-BIXEPS 8", "Post-BIXEPS 8", "symptom", 1.0, "Wellbeing"),
    Metric("Health Score", "Pre-BIXEPS 9", "Post-BIXEPS 9", "higher", 1.5, "Wellbeing"),
]

IDENTITY_HEADERS = [
    "S/N", "Name", "Gender", "Age (in 2026)", "Date of Pre-BIXEPS assessment",
    "Frailty Status", "Frailty Status 2",
]

OUTCOME_FILLS = {
    "Improved": "C6EFCE",
    "Maintained": "DDEBF7",
    "No Change": "E7E6E6",
    "Declined": "FFC7CE",
    "Insufficient Data": "FFF2CC",
    "Missing Data": "FFF2CC",
}


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def find_header_row(ws) -> int:
    required = {clean_header("Name"), clean_header("SPPB Overall Score")}
    for row in range(1, min(ws.max_row, 30) + 1):
        values = {clean_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if required.issubset(values):
            return row
    raise ValueError("Could not find the header row. The sheet must contain 'Name' and 'SPPB Overall Score'.")


def header_map(ws, header_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = clean_header(ws.cell(header_row, col).value)
        if key and key not in result:
            result[key] = col
    return result


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip().casefold() in {"", "na", "n/a", "nil", "-", "not available"}:
        return True
    return False


def to_number(value: Any) -> float | None:
    if is_missing(value):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else None


def symptom_severity(value: Any) -> float | None:
    if is_missing(value):
        return None
    numeric = to_number(value)
    if numeric is not None and not isinstance(value, str):
        return numeric
    text = str(value).strip().casefold()
    if text in {"no", "none", "never", "false", "0"}:
        return 0.0
    if any(term in text for term in ["once in a while", "occasionally", "sometimes", "mild", "1-2", "1 to 2"]):
        return 1.0
    if text in {"yes", "true"} or any(term in text for term in ["3-4", "3 to 4", "often", "frequent", "weekly", "month"]):
        return 2.0
    if any(term in text for term in ["daily", "every night", "severe", "always"]):
        return 3.0
    if text.startswith("yes"):
        return 2.0
    numeric_text = to_number(text)
    return numeric_text


def compare(pre: Any, post: Any, direction: str) -> tuple[str, float | None]:
    parser: Callable[[Any], float | None] = symptom_severity if direction == "symptom" else to_number
    pre_num, post_num = parser(pre), parser(post)
    if pre_num is None or post_num is None:
        return "Missing Data", None
    delta = post_num - pre_num
    if abs(delta) < 1e-9:
        return "No Change", 0.0
    improved = delta > 0 if direction == "higher" else delta < 0
    return ("Improved" if improved else "Declined"), delta


def normalise_frailty(value: Any) -> int | None:
    if is_missing(value):
        return None
    text = str(value).strip().casefold()
    number = to_number(value)
    if number is not None and text.replace(".", "", 1).isdigit():
        return int(number)
    # Lower rank is better.
    labels = [
        (0, ["robust", "fit", "not frail", "normal"]),
        (1, ["pre-frail", "pre frail", "prefrail", "vulnerable"]),
        (2, ["frail", "mild frailty", "mildly frail"]),
        (3, ["moderate", "moderately frail"]),
        (4, ["severe", "severely frail"]),
    ]
    for rank, terms in labels:
        if any(term in text for term in terms):
            return rank
    return None


def compare_frailty(pre: Any, post: Any) -> str:
    p, q = normalise_frailty(pre), normalise_frailty(post)
    if p is None or q is None:
        return "Missing Data"
    if q < p:
        return "Improved"
    if q > p:
        return "Declined"
    return "No Change"


def frailty_transition(pre: Any, post: Any) -> tuple[str, bool]:
    """Return a readable frailty transition and whether a frail senior became normal/robust."""
    p, q = normalise_frailty(pre), normalise_frailty(post)
    if p is None or q is None:
        return "Missing Data", False
    pre_text = str(pre).strip().title()
    post_text = str(post).strip().title()
    transition = f"{pre_text} → {post_text}"
    recovered = p >= 2 and q == 0
    return transition, recovered


def _fmt_value(value: Any) -> str:
    number = to_number(value)
    if number is not None:
        return f"{number:g}"
    return str(value).strip() if not is_missing(value) else "missing"


def metric_finding(metric: Metric, pre: Any, post: Any, status: str) -> str:
    """Create a plain-language reason for one pre/post finding."""
    if status == "Missing Data":
        return f"{metric.label}: incomplete pre/post data"
    pre_text, post_text = _fmt_value(pre), _fmt_value(post)
    if status == "No Change":
        return f"{metric.label} was unchanged at {post_text}"
    if metric.direction == "higher":
        verb = "increased" if status == "Improved" else "decreased"
    else:
        verb = "decreased" if status == "Improved" else "increased"
    return f"{metric.label} {verb} from {pre_text} to {post_text}"


def domain_reason(record: dict[str, Any], outcomes: list[tuple[Metric, str]], domain: str) -> str:
    findings: list[str] = []
    for metric, status in outcomes:
        if metric.domain != domain or status == "Missing Data":
            continue
        findings.append(metric_finding(metric, record.get(f"{metric.label} Pre"), record.get(f"{metric.label} Post"), status))
    if not findings:
        return "Insufficient comparable pre/post data."
    return "; ".join(findings) + "."


def detailed_analysis_reason(record: dict[str, Any], outcomes: list[tuple[Metric, str]]) -> str:
    """Explain why the senior received the overall classification using exact findings."""
    overall = record.get("Overall Outcome", "")
    improved = [(m, s) for m, s in outcomes if s == "Improved"]
    declined = [(m, s) for m, s in outcomes if s == "Declined"]
    unchanged = sum(s == "No Change" for _, s in outcomes)
    comparable = sum(s != "Missing Data" for _, s in outcomes)

    if overall == "Insufficient Data":
        return f"Insufficient data because only {comparable}/{len(METRICS)} assessment areas had complete pre/post results."
    if overall == "No Change":
        return f"No change because all {comparable} comparable assessment areas remained the same."

    lead = {
        "Improved": "Classified as Improved because positive changes outweighed declines.",
        "Declined": "Classified as Declined because negative changes outweighed improvements.",
        "Maintained": "Classified as Maintained because the results were stable or mixed without a clear overall shift.",
    }.get(overall, f"Classified as {overall}.")

    parts = [lead]
    transition = record.get("Frailty Transition")
    if record.get("Recovered from Frailty"):
        parts.append(f"Frailty status improved from {transition}; this is a key positive programme outcome.")
    elif record.get("Frailty Outcome") == "Improved":
        parts.append(f"Frailty status improved from {transition}.")
    elif record.get("Frailty Outcome") == "Declined":
        parts.append(f"Frailty status worsened from {transition}.")

    if improved:
        texts = [metric_finding(m, record.get(f"{m.label} Pre"), record.get(f"{m.label} Post"), s) for m, s in improved[:5]]
        parts.append("Improvements: " + "; ".join(texts) + ".")
    if declined:
        texts = [metric_finding(m, record.get(f"{m.label} Pre"), record.get(f"{m.label} Post"), s) for m, s in declined[:5]]
        parts.append("Declines: " + "; ".join(texts) + ".")
    parts.append(
        f"Overall, {len(improved)} area(s) improved, {len(declined)} declined and {unchanged} were unchanged; "
        f"{comparable}/{len(METRICS)} areas were comparable."
    )
    return " ".join(parts)


def domain_result(outcomes: list[tuple[Metric, str]], domain: str) -> str:
    selected = [(m, s) for m, s in outcomes if m.domain == domain and s != "Missing Data"]
    if not selected:
        return "Insufficient Data"
    improved = sum(m.weight for m, s in selected if s == "Improved")
    declined = sum(m.weight for m, s in selected if s == "Declined")
    unchanged = sum(s == "No Change" for _, s in selected)
    if unchanged == len(selected):
        return "No Change"
    if improved >= declined + 1.0:
        return "Improved"
    if declined >= improved + 1.0:
        return "Declined"
    return "Maintained"


def overall_result(outcomes: list[tuple[Metric, str]]) -> tuple[str, str, float, float]:
    available = [(metric, status) for metric, status in outcomes if status != "Missing Data"]
    if not available:
        return "Insufficient Data", "No complete pre/post comparisons were available.", 0.0, 0.0

    improved_weight = sum(metric.weight for metric, status in available if status == "Improved")
    declined_weight = sum(metric.weight for metric, status in available if status == "Declined")
    unchanged = sum(1 for _, status in available if status == "No Change")
    improved = sum(1 for _, status in available if status == "Improved")
    declined = sum(1 for _, status in available if status == "Declined")
    completeness = len(available) / len(METRICS)
    improvement_score = improved_weight - declined_weight

    if completeness < 0.25:
        result = "Insufficient Data"
    elif unchanged == len(available):
        result = "No Change"
    elif improved_weight >= declined_weight + 1.5 and improved >= 2:
        result = "Improved"
    elif declined_weight >= improved_weight + 1.5 and declined >= 2:
        result = "Declined"
    else:
        result = "Maintained"

    reason = (
        f"{improved} area(s) improved, {declined} declined and {unchanged} had no change; "
        f"{len(available)}/{len(METRICS)} areas were comparable."
    )
    return result, reason, completeness, improvement_score


def follow_up_reason(record: dict[str, Any]) -> str:
    reasons: list[str] = []
    if record.get("Overall Outcome") == "Declined":
        reasons.append("overall assessment declined")
    if record.get("SPPB Outcome") == "Declined":
        reasons.append("SPPB worsened")
    if record.get("Mobility Outcome") == "Declined":
        reasons.append("mobility worsened")
    if record.get("Overall Pain Outcome") == "Declined":
        reasons.append("overall pain increased")
    if record.get("Frailty Outcome") == "Declined":
        reasons.append("frailty status worsened")
    if record.get("Data Completeness", 0) < 0.25:
        reasons.append("post-assessment data incomplete")
    return "; ".join(reasons).capitalize() + ("." if reasons else "")


def _style_sheet(ws, headers: list[str]) -> None:
    dark = "1F4E78"
    thin = Side(style="thin", color="D9E1F2")
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 36

    status_columns = [i + 1 for i, h in enumerate(headers) if h.endswith("Outcome") or h.endswith("Domain")]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in status_columns:
            cell = ws.cell(row[0].row, col)
            if cell.value in OUTCOME_FILLS:
                cell.fill = PatternFill("solid", fgColor=OUTCOME_FILLS[cell.value])

    for idx, header in enumerate(headers, start=1):
        if header == "Name":
            width = 24
        elif header in {"Analysis Reason", "Follow-up Reason"}:
            width = 50
        elif "Outcome" in header or "Domain" in header:
            width = 18
        elif header.endswith("Pre") or header.endswith("Post") or header.endswith("Change"):
            width = 15
        else:
            width = min(max(len(header) + 2, 12), 22)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_dataframe_sheet(wb, title: str, df: pd.DataFrame) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    _style_sheet(ws, headers)
    if "Data Completeness" in headers:
        col = headers.index("Data Completeness") + 1
        for cell in ws[get_column_letter(col)][1:]:
            cell.number_format = "0%"


def analyse_workbook(uploaded_bytes: bytes, selected_sheet: str | None = None):
    values_wb = load_workbook(io.BytesIO(uploaded_bytes), data_only=True)
    source_wb = load_workbook(io.BytesIO(uploaded_bytes))

    sheet_name = selected_sheet or values_wb.sheetnames[0]
    values_ws = values_wb[sheet_name]
    header_row = find_header_row(values_ws)
    columns = header_map(values_ws, header_row)

    missing_headers = []
    for metric in METRICS:
        if clean_header(metric.pre_header) not in columns or clean_header(metric.post_header) not in columns:
            missing_headers.append(f"{metric.pre_header} / {metric.post_header}")
    if clean_header("Name") not in columns:
        missing_headers.append("Name")
    if missing_headers:
        raise ValueError("Missing required fields: " + "; ".join(missing_headers))

    records: list[dict[str, Any]] = []
    for row in range(header_row + 1, values_ws.max_row + 1):
        name = values_ws.cell(row, columns[clean_header("Name")]).value
        if is_missing(name):
            continue
        record: dict[str, Any] = {}
        for field in IDENTITY_HEADERS:
            col = columns.get(clean_header(field))
            record[field] = values_ws.cell(row, col).value if col else None

        outcomes: list[tuple[Metric, str]] = []
        for metric in METRICS:
            pre = values_ws.cell(row, columns[clean_header(metric.pre_header)]).value
            post = values_ws.cell(row, columns[clean_header(metric.post_header)]).value
            status, delta = compare(pre, post, metric.direction)
            record[f"{metric.label} Pre"] = pre
            record[f"{metric.label} Post"] = post
            record[f"{metric.label} Change"] = delta
            record[f"{metric.label} Outcome"] = status
            outcomes.append((metric, status))

        record["Frailty Outcome"] = compare_frailty(record.get("Frailty Status"), record.get("Frailty Status 2"))
        transition, recovered = frailty_transition(record.get("Frailty Status"), record.get("Frailty Status 2"))
        record["Frailty Transition"] = transition
        record["Recovered from Frailty"] = recovered
        record["Frailty Evidence"] = (
            "Verified with complete SPPB" if recovered and record.get("SPPB Change") is not None
            else "Limited: SPPB pre/post incomplete" if recovered
            else "Not applicable"
        )
        for domain in ["Physical Function", "Pain", "Daily Living", "Wellbeing"]:
            record[f"{domain} Domain"] = domain_result(outcomes, domain)
            record[f"{domain} Reason"] = domain_reason(record, outcomes, domain)

        result, reason, completeness, score = overall_result(outcomes)
        record["Improved Areas"] = sum(status == "Improved" for _, status in outcomes)
        record["Declined Areas"] = sum(status == "Declined" for _, status in outcomes)
        record["No-Change Areas"] = sum(status == "No Change" for _, status in outcomes)
        record["Comparable Areas"] = sum(status != "Missing Data" for _, status in outcomes)
        record["Data Completeness"] = completeness
        record["Improvement Score"] = score
        record["Overall Outcome"] = result
        record["Analysis Reason"] = detailed_analysis_reason(record, outcomes)
        record["Critical Decline"] = any(
            record.get(field) == "Declined"
            for field in ["SPPB Outcome", "Mobility Outcome", "Overall Pain Outcome", "Frailty Outcome"]
        )
        record["Follow-up Reason"] = follow_up_reason(record)
        records.append(record)

    df = pd.DataFrame(records)
    if df.empty:
        raise ValueError("No senior records were found below the header row.")

    for old_name in ["Senior Analysis", "Analysis Dashboard", "Follow-up List", "Top Improvers", "Frailty Improvements"]:
        if old_name in source_wb.sheetnames:
            del source_wb[old_name]

    _write_dataframe_sheet(source_wb, "Senior Analysis", df)

    follow_cols = [
        "Name", "Gender", "Age (in 2026)", "Overall Outcome", "SPPB Outcome", "Mobility Outcome",
        "Overall Pain Outcome", "Frailty Outcome", "Follow-up Reason",
    ]
    follow_df = df[(df["Overall Outcome"] == "Declined") | df["Critical Decline"]][follow_cols]
    _write_dataframe_sheet(source_wb, "Follow-up List", follow_df)

    top_cols = [
        "Name", "Overall Outcome", "Improvement Score", "Improved Areas", "Declined Areas",
        "SPPB Change", "Overall Pain Change", "Health Score Change", "Analysis Reason",
    ]
    top_df = df[df["Overall Outcome"] != "Insufficient Data"].sort_values(
        ["Improvement Score", "Improved Areas", "SPPB Change"], ascending=[False, False, False]
    ).head(10)[top_cols]
    _write_dataframe_sheet(source_wb, "Top Improvers", top_df)

    frailty_cols = [
        "Name", "Gender", "Age (in 2026)", "Frailty Status", "Frailty Status 2", "Frailty Transition", "Frailty Evidence",
        "SPPB Pre", "SPPB Post", "SPPB Change", "Overall Outcome", "Analysis Reason",
    ]
    frailty_df = df[df["Recovered from Frailty"]][frailty_cols].copy()
    _write_dataframe_sheet(source_wb, "Frailty Improvements", frailty_df)

    dashboard = source_wb.create_sheet("Analysis Dashboard", 0)
    dark = "1F4E78"
    dashboard["A1"] = "BIXEPS Senior Outcome Dashboard"
    dashboard["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    dashboard["A1"].fill = PatternFill("solid", fgColor=dark)
    dashboard.merge_cells("A1:D1")
    dashboard.append([])
    dashboard.append(["Outcome", "Seniors", "Percentage", "Definition"])
    for cell in dashboard[3]:
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(color="FFFFFF", bold=True)

    definitions = {
        "Improved": "Clear positive change across multiple weighted areas.",
        "Maintained": "Stable or mixed results without clear overall decline.",
        "No Change": "All comparable pre/post results were identical.",
        "Declined": "Clear negative change across multiple weighted areas.",
        "Insufficient Data": "Too few complete pre/post comparisons.",
    }
    counts = df["Overall Outcome"].value_counts()
    outcomes_order = ["Improved", "Maintained", "No Change", "Declined", "Insufficient Data"]
    for row_num, outcome in enumerate(outcomes_order, start=4):
        count = int(counts.get(outcome, 0))
        dashboard.cell(row_num, 1, outcome)
        dashboard.cell(row_num, 2, count)
        dashboard.cell(row_num, 3, count / len(df))
        dashboard.cell(row_num, 3).number_format = "0.0%"
        dashboard.cell(row_num, 4, definitions[outcome])
        dashboard.cell(row_num, 1).fill = PatternFill("solid", fgColor=OUTCOME_FILLS[outcome])

    complete = df[df["Overall Outcome"] != "Insufficient Data"]
    dashboard["A11"] = "Total seniors analysed"
    dashboard["B11"] = len(df)
    dashboard["A12"] = "Average data completeness"
    dashboard["B12"] = float(df["Data Completeness"].mean())
    dashboard["B12"].number_format = "0.0%"
    dashboard["A13"] = "Average SPPB change"
    dashboard["B13"] = float(complete["SPPB Change"].mean()) if not complete.empty else None
    dashboard["A14"] = "Average overall pain change"
    dashboard["B14"] = float(complete["Overall Pain Change"].mean()) if not complete.empty else None
    dashboard["A15"] = "Average health-score change"
    dashboard["B15"] = float(complete["Health Score Change"].mean()) if not complete.empty else None
    dashboard["A16"] = "Seniors flagged for follow-up"
    dashboard["B16"] = len(follow_df)
    pre_frail = df[df["Frailty Status"].apply(normalise_frailty).ge(2)]
    pre_frail_complete = pre_frail[pre_frail["Frailty Status 2"].apply(normalise_frailty).notna()]
    recovered_count = int(df["Recovered from Frailty"].sum())
    verified_recovered = int(((df["Recovered from Frailty"]) & df["SPPB Change"].notna()).sum())
    dashboard["A17"] = "Verified frail to normal"
    dashboard["B17"] = verified_recovered
    dashboard["C17"] = verified_recovered / len(pre_frail_complete) if len(pre_frail_complete) else None
    dashboard["C17"].number_format = "0.0%"

    dashboard["A19"] = "Classification note"
    dashboard["A20"] = (
        "SPPB and Health Score: higher is better. Pain, daily-function and questionnaire scores: lower is better. "
        "Symptom text is converted to a simple severity scale. Results with fewer than 25% comparable areas are "
        "marked Insufficient Data. Follow-up flags include overall decline or worsening in SPPB, mobility, pain or frailty."
    )
    dashboard.merge_cells("A20:D23")
    dashboard["A20"].alignment = Alignment(wrap_text=True, vertical="top")
    dashboard.column_dimensions["A"].width = 28
    dashboard.column_dimensions["B"].width = 16
    dashboard.column_dimensions["C"].width = 14
    dashboard.column_dimensions["D"].width = 58
    dashboard.freeze_panes = "A3"

    chart = BarChart()
    chart.title = "Senior Outcomes"
    chart.y_axis.title = "Number of Seniors"
    chart.x_axis.title = "Outcome"
    data = Reference(dashboard, min_col=2, min_row=3, max_row=8)
    categories = Reference(dashboard, min_col=1, min_row=4, max_row=8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 13
    dashboard.add_chart(chart, "F3")

    output = io.BytesIO()
    source_wb.save(output)
    output.seek(0)
    return df, output.getvalue(), sheet_name


def build_pdf_report(df: pd.DataFrame, report_name: str = "BIXEPS") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"{report_name} Management Report",
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="CenteredTitle", parent=styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#1F4E78")))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=8, leading=10))
    story = [Paragraph("BIXEPS Management Report", styles["CenteredTitle"]), Spacer(1, 6)]

    counts = df["Overall Outcome"].value_counts()
    outcome_order = ["Improved", "Maintained", "No Change", "Declined", "Insufficient Data"]
    overview = [["Outcome", "Seniors", "Percentage"]]
    for outcome in outcome_order:
        count = int(counts.get(outcome, 0))
        overview.append([outcome, count, f"{count / len(df):.1%}"])
    table = Table(overview, colWidths=[55 * mm, 35 * mm, 35 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E1F2")),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FB")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.extend([table, Spacer(1, 10)])

    complete = df[df["Overall Outcome"] != "Insufficient Data"]
    key = [
        ["Key indicator", "Result"],
        ["Seniors analysed", str(len(df))],
        ["Average SPPB change", f"{complete['SPPB Change'].mean():+.2f}" if not complete.empty else "N/A"],
        ["Average overall pain change", f"{complete['Overall Pain Change'].mean():+.2f}" if not complete.empty else "N/A"],
        ["Average health-score change", f"{complete['Health Score Change'].mean():+.2f}" if not complete.empty else "N/A"],
        ["Seniors requiring review", str(int(((df['Overall Outcome'] == 'Declined') | df['Critical Decline']).sum()))],
        ["Verified frail to normal", str(int(((df["Recovered from Frailty"]) & df["SPPB Change"].notna()).sum()))],
    ]
    key_table = Table(key, colWidths=[80 * mm, 45 * mm])
    key_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E1F2")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FB")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.extend([key_table, PageBreak()])

    story.append(Paragraph("Seniors Requiring Follow-up", styles["Heading1"]))
    follow = df[(df["Overall Outcome"] == "Declined") | df["Critical Decline"]].copy()
    follow_data = [["Name", "Overall", "SPPB", "Mobility", "Pain", "Frailty", "Reason"]]
    if follow.empty:
        follow_data.append(["None", "", "", "", "", "", "No seniors met the follow-up criteria."])
    else:
        for _, row in follow.head(30).iterrows():
            follow_data.append([
                str(row["Name"]), str(row["Overall Outcome"]), str(row["SPPB Outcome"]),
                str(row["Mobility Outcome"]), str(row["Overall Pain Outcome"]), str(row["Frailty Outcome"]),
                Paragraph(str(row["Follow-up Reason"]), styles["Small"]),
            ])
    follow_table = Table(follow_data, repeatRows=1, colWidths=[38*mm, 24*mm, 24*mm, 24*mm, 24*mm, 24*mm, 90*mm])
    follow_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E1F2")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FB")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(follow_table)

    story.append(PageBreak())
    story.append(Paragraph("Top Improving Seniors", styles["Heading1"]))
    top = df[df["Overall Outcome"] != "Insufficient Data"].sort_values(
        ["Improvement Score", "Improved Areas", "SPPB Change"], ascending=[False, False, False]
    ).head(10)
    top_data = [["Name", "Overall", "Score", "Improved Areas", "Declined Areas", "SPPB Change", "Pain Change", "Health Change"]]
    for _, row in top.iterrows():
        top_data.append([
            str(row["Name"]), str(row["Overall Outcome"]), f"{row['Improvement Score']:.1f}",
            int(row["Improved Areas"]), int(row["Declined Areas"]),
            "" if pd.isna(row["SPPB Change"]) else f"{row['SPPB Change']:+.1f}",
            "" if pd.isna(row["Overall Pain Change"]) else f"{row['Overall Pain Change']:+.1f}",
            "" if pd.isna(row["Health Score Change"]) else f"{row['Health Score Change']:+.1f}",
        ])
    top_table = Table(top_data, repeatRows=1, colWidths=[48*mm, 30*mm, 22*mm, 30*mm, 30*mm, 28*mm, 28*mm, 28*mm])
    top_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E1F2")),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FB")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(top_table)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "Interpretation note: higher SPPB and health scores are positive. Lower pain, daily-function problem and questionnaire scores are positive. "
        "The automated classification supports review but should not replace clinical or professional judgement.",
        styles["Small"],
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
